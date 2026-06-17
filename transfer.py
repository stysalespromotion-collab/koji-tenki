# v2
"""
転記ロジックモジュール
見積決定（.xlsx / .xls）→ 出来高.xlsx / 工事完了.xlsx
"""

import io
from openpyxl import load_workbook


def _load_workbook_any(file_bytes, filename):
    fname = filename.lower()
    if fname.endswith('.xls') and not fname.endswith('.xlsx'):
        import xlrd
        from openpyxl import Workbook as OpenpyxlWorkbook
        xls_wb = xlrd.open_workbook(file_contents=file_bytes)
        wb = OpenpyxlWorkbook()
        wb.remove(wb.active)
        for sheet_name in xls_wb.sheet_names():
            xls_ws = xls_wb.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            for row in range(xls_ws.nrows):
                for col in range(xls_ws.ncols):
                    ws.cell(row + 1, col + 1, xls_ws.cell(row, col).value)
        return wb
    else:
        return load_workbook(io.BytesIO(file_bytes), data_only=True)


def _set_val(ws, coord, value):
    if value is not None and str(value).strip() not in ('', 'None'):
        ws[coord] = value


def _find_uriage_gokei(ws_src):
    for row in range(10, 30):
        b_val = ws_src.cell(row, 2).value
        if b_val and ('合' in str(b_val) and '計' in str(b_val)):
            val = ws_src.cell(row, 8).value
            if val and isinstance(val, (int, float)) and val > 0:
                return val
    return None


def _find_shitauke(ws_src):
    results = []
    current_name = None
    for row in range(19, 55):
        b_val = ws_src.cell(row, 2).value
        h_val = ws_src.cell(row, 8).value
        e_val = ws_src.cell(row, 5).value
        g_val = ws_src.cell(row, 7).value
        if b_val is None:
            continue
        b_str = str(b_val).strip()
        if b_str == '':
            continue
        if any(kw in b_str for kw in ['下請工事店名', '取引条件', '出精値引き', '材料']):
            continue
        if '下請支払合計' in b_str:
            break
        if '合' in b_str and '計' in b_str:
            if current_name and h_val is not None and isinstance(h_val, (int, float)):
                results.append((current_name, h_val))
                current_name = None
            continue
        e_empty = (e_val is None or str(e_val).strip() == '')
        g_empty = (g_val is None or str(g_val).strip() == '')
        if e_empty and g_empty:
            current_name = b_str
    return results


def _find_material(ws_src):
    results = []
    mat_start = None
    for row in range(35, 65):
        b_val = ws_src.cell(row, 2).value
        if b_val and '材料' in str(b_val) and ('販売' in str(b_val) or '使用' in str(b_val)):
            mat_start = row + 1
            break
    if mat_start is None:
        return results
    for row in range(mat_start, mat_start + 20):
        b_val = ws_src.cell(row, 2).value
        if b_val is None or str(b_val).strip() == '':
            continue
        b_str = str(b_val).strip()
        if '合' in b_str and '計' in b_str:
            break
        if any(kw in b_str for kw in ['施工担当者', '下請業者', 'その他経費', '間接経費']):
            break
        row_data = {}
        for col in [2, 4, 5, 6, 7, 8, 9]:
            val = ws_src.cell(row, col).value
            if val is not None and str(val).strip() not in ('', 'None'):
                row_data[col] = val
        if row_data:
            results.append(row_data)
    return results


def _find_toko_mgmt(ws_src):
    result = {}
    for row in range(50, 70):
        i_val = ws_src.cell(row, 9).value
        if i_val is None:
            continue
        i_str = str(i_val).strip()
        if '施工面積' in i_str:
            result['menseki'] = (ws_src.cell(row, 10).value, ws_src.cell(row, 11).value)
        elif '塗厚' in i_str:
            result['toko'] = (ws_src.cell(row, 10).value, ws_src.cell(row, 11).value)
        elif '袋容量' in i_str:
            result['fukuro'] = (ws_src.cell(row, 10).value, ws_src.cell(row, 11).value)
    return result


def transfer_dekidaka(mitsumori_bytes, mitsumori_name, kaime):
    wb_src = _load_workbook_any(mitsumori_bytes, mitsumori_name)
    ws_src = wb_src['見積決定']

    with open('templates/出来高.xlsx', 'rb') as f:
        tmpl_bytes = f.read()
    wb_dst = load_workbook(io.BytesIO(tmpl_bytes))
    ws_dst = wb_dst['出来高']

    _set_val(ws_dst, 'B3', ws_src['B3'].value)
    _set_val(ws_dst, 'L2', ws_src['K2'].value)
    _set_val(ws_dst, 'L3', ws_src['K3'].value)
    _set_val(ws_dst, 'L4', ws_src['K4'].value)
    _set_val(ws_dst, 'C6', ws_src['C6'].value)
    _set_val(ws_dst, 'C7', ws_src['C7'].value)
    _set_val(ws_dst, 'C8', ws_src['C8'].value)
    _set_val(ws_dst, 'I6', ws_src['I6'].value)
    _set_val(ws_dst, 'I7', ws_src['I7'].value)
    _set_val(ws_dst, 'I8', ws_src['I8'].value)

    ws_dst['J10'] = f'{kaime}'
    ws_dst['J11'] = f'{kaime}'

    uriage = _find_uriage_gokei(ws_src)
    if uriage:
        ws_dst['F9'] = uriage

    shitauke = _find_shitauke(ws_src)
    for i, (name, amount) in enumerate(shitauke):
        dst_row = 17 + i
        if dst_row > 20:
            break
        ws_dst.cell(dst_row, 2).value = name
        ws_dst.cell(dst_row, 8).value = amount

    materials = _find_material(ws_src)
    for i, mat in enumerate(materials):
        dst_row = 23 + i
        if dst_row > 37:
            break
        for col, val in mat.items():
            ws_dst.cell(dst_row, col).value = val

    out = io.BytesIO()
    wb_dst.save(out)
    out.seek(0)
    return out.read()


def transfer_kanryo(mitsumori_bytes, mitsumori_name, dekidaka_list, kaime):
    wb_src = _load_workbook_any(mitsumori_bytes, mitsumori_name)
    ws_src = wb_src['見積決定']

    with open('templates/工事完了.xlsx', 'rb') as f:
        tmpl_bytes = f.read()
    wb_dst = load_workbook(io.BytesIO(tmpl_bytes))
    ws_dst = wb_dst['１回完了']

    _set_val(ws_dst, 'A2', ws_src['B3'].value)
    _set_val(ws_dst, 'C3', ws_src['C4'].value)
    _set_val(ws_dst, 'C5', ws_src['C6'].value)
    _set_val(ws_dst, 'C6', ws_src['C7'].value)
    _set_val(ws_dst, 'C7', ws_src['C8'].value)
    _set_val(ws_dst, 'I5', ws_src['I6'].value)
    _set_val(ws_dst, 'I6', ws_src['I7'].value)
    _set_val(ws_dst, 'I7', ws_src['I8'].value)
    _set_val(ws_dst, 'K1', ws_src['K2'].value)
    _set_val(ws_dst, 'K2', ws_src['K3'].value)
    _set_val(ws_dst, 'K3', ws_src['K4'].value)

    uriage = _find_uriage_gokei(ws_src)
    if uriage:
        ws_dst['F8'] = uriage

    shita_totals = {}
    shita_order = []
    for db, dname in dekidaka_list:
        wb_d = _load_workbook_any(db, dname)
        ws_d = wb_d['出来高']
        for r in range(17, 21):
            name = ws_d.cell(r, 2).value
            amount = ws_d.cell(r, 8).value
            if name and str(name).strip():
                name = str(name).strip()
                if name not in shita_totals:
                    shita_totals[name] = 0
                    shita_order.append(name)
                shita_totals[name] += (amount or 0)

    for i, name in enumerate(shita_order):
        dst_row = 17 + i
        if dst_row > 22:
            break
        ws_dst.cell(dst_row, 2).value = name
        ws_dst.cell(dst_row, 8).value = shita_totals[name]

    materials = _find_material(ws_src)
    for i, mat in enumerate(materials):
        dst_row = 25 + i
        if dst_row > 39:
            break
        for col, val in mat.items():
            ws_dst.cell(dst_row, col).value = val

    toko = _find_toko_mgmt(ws_src)
    if 'menseki' in toko:
        j, k = toko['menseki']
        if j is not None: ws_dst.cell(54, 10).value = j
        if k is not None: ws_dst.cell(54, 11).value = k
    if 'toko' in toko:
        j, k = toko['toko']
        if j is not None: ws_dst.cell(55, 10).value = j
        if k is not None: ws_dst.cell(55, 11).value = k
    if 'fukuro' in toko:
        j, k = toko['fukuro']
        if j is not None: ws_dst.cell(56, 10).value = j
        if k is not None: ws_dst.cell(56, 11).value = k

    out = io.BytesIO()
    wb_dst.save(out)
    out.seek(0)
    return out.read()
